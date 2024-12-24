import ambr
from openpyxl.styles import (
    NamedStyle,
    Alignment,
    Font,
    Border,
    Side,
    PatternFill
)
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from api import API

api = API()
category = input("Набор достижений: ")
raw_achievements = api.get_achievements_by_category(category)
raw_achievements.sort(key=lambda a: a.order)

result_achievements = list()
achievement: ambr.Achievement
for achievement in raw_achievements:
    for detail in achievement.details:
        result_achievements.append({
            "Название достижения": detail.title,
            "Описание достижения": detail.description,
            "올가": " ",
            "미샤": " "
        })

workbook: Workbook = Workbook()
worksheet: Worksheet = workbook.active

header = list(result_achievements[0].keys())
worksheet.append(header)
for result in result_achievements:
    worksheet.append([result[column] for column in header])

border_style = 'medium'
completed_color = 'e1b4ca'
header_color = '000000'
cell_color = 'd4d4d4'
fill_color = 'ffcbe5'

completed_style = NamedStyle(name="Выполнено")
completed_style.font = Font(sz=14)
completed_style.alignment = Alignment(
    vertical="center",
    wrap_text=True
)
completed_style.fill = PatternFill(
    start_color=fill_color,
    end_color=fill_color,
    fill_type='solid'
)
completed_border_style = Border(
    left=Side(border_style=border_style, color=completed_color),
    right=Side(border_style=border_style, color=completed_color),
    top=Side(border_style=border_style, color=completed_color),
    bottom=Side(border_style=border_style, color=completed_color)
)
completed_style.border = completed_border_style
for cell in worksheet[1]:
    cell.style = completed_style

header_style = NamedStyle(name="Заголовок")
header_style.font = Font(sz=18, bold=True)
header_style.alignment = Alignment(horizontal="center")
header_border_style = Border(
    left=Side(border_style=border_style, color=header_color),
    right=Side(border_style=border_style, color=header_color),
    top=Side(border_style=border_style, color=header_color),
    bottom=Side(border_style=border_style, color=header_color)
)
header_style.border = header_border_style
for cell in worksheet[1]:
    cell.style = header_style

cell_style = NamedStyle(name="Ячейка")
cell_style.font = Font(sz=14)
cell_style.alignment = Alignment(
    vertical="center",
    wrap_text=True
)
cell_border_style = Border(
    left=Side(border_style=border_style, color=cell_color),
    right=Side(border_style=border_style, color=cell_color),
    top=Side(border_style=border_style, color=cell_color),
    bottom=Side(border_style=border_style, color=cell_color)
)
cell_style.border = cell_border_style
for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column
):
    for cell in row:
        cell.style = cell_style

worksheet.column_dimensions['A'].width = 70
worksheet.column_dimensions['B'].width = 70
worksheet.column_dimensions['C'].width = 15
worksheet.column_dimensions['D'].width = 15

workbook.save("achievements.xlsx")
